"""报告生成器——将评测结果格式化为可视化报告"""

import json
import datetime
from pathlib import Path
from io import BytesIO

from config import settings
from core.evaluator import EvalResult, DIMENSIONS, DimensionScore
from core.dialogue_runner import DialogResult
from core.scenario_builder import TaskRubric


class ReportGenerator:
    """评测报告生成器"""

    def generate_markdown(
        self,
        dialog_result: DialogResult,
        eval_result: EvalResult,
        rubric: TaskRubric,
        task_instruction: str,
    ) -> str:
        """生成 Markdown 格式评测报告"""
        lines = []

        lines.append("# EvalAgent 多轮对话评测报告\n")
        lines.append(f"**生成时间**: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        # 一、任务概述
        lines.append("## 一、任务概述\n")
        lines.append(f"**任务指令**: {task_instruction}\n")
        lines.append(f"**任务目标**: {rubric.task_goal}\n")
        if rubric.role:
            lines.append(f"**AI 角色**: {rubric.role}\n")
        if rubric.must_do:
            lines.append(f"**必须完成**: {'、'.join(rubric.must_do)}\n")
        if rubric.must_not_do:
            lines.append(f"**禁止行为**: {'、'.join(rubric.must_not_do)}\n")
        if rubric.opening_line:
            lines.append(f"**开场白模板**: {rubric.opening_line}\n")
        if rubric.call_flow:
            lines.append(f"\n### 指定通话流程\n")
            for step in rubric.call_flow:
                lines.append(f"- **步骤{step.step_id}**: {step.title}")
                if step.reference_script:
                    lines.append(f"  （参考话术: {step.reference_script[:60]}...）\n")
                else:
                    lines.append("\n")
        if rubric.knowledge_points:
            lines.append(f"\n### 知识库 (FAQ)\n")
            for q, a in rubric.knowledge_points.items():
                lines.append(f"- **{q}**: {a[:80]}...\n")

        # 二、评测场景
        lines.append("## 二、评测场景\n")
        s = dialog_result.scenario
        lines.append(f"**用户画像**: {s.persona_name}\n")
        lines.append(f"**测试目标**: {s.test_goal}\n")
        lines.append(f"**结束原因**: {self._end_reason_text(dialog_result.finished_reason)}\n")
        lines.append(f"**实际轮次**: {len(dialog_result.turns)} 轮\n")

        # 三、综合评分
        lines.append("## 三、综合评分\n")
        lines.append(f"### 总分: **{eval_result.overall_score}/100**\n")
        if eval_result.overall_score >= 80:
            lines.append("> 🟢 优秀 — 模型表现良好\n")
        elif eval_result.overall_score >= 60:
            lines.append("> 🟡 一般 — 存在可改进空间\n")
        else:
            lines.append("> 🔴 较差 — 需要重点关注\n")

        # 四、各维度评分
        lines.append("## 四、各维度评分\n")
        lines.append("| 维度 | 权重 | 得分 | 评分理由 |")
        lines.append("|------|------|------|----------|")
        for dim in DIMENSIONS:
            key = dim["key"]
            if key in eval_result.dimensions:
                ds = eval_result.dimensions[key]
                score_bar = "█" * ds.score + "░" * (5 - ds.score)
                lines.append(
                    f"| {dim['name']} | {dim['weight']*100:.0f}% "
                    f"| {ds.score}/5 {score_bar} | {ds.reason} |"
                )

        # 五、违规项
        if eval_result.violations:
            lines.append("\n## 五、违规项\n")
            for v in eval_result.violations:
                lines.append(f"- ⚠️ {v}")

        # 六、亮点
        if eval_result.good_points:
            lines.append("\n## 六、表现亮点\n")
            for p in eval_result.good_points:
                lines.append(f"- ✅ {p}")

        # 七、对话记录
        lines.append("\n## 七、对话记录\n")
        for turn in dialog_result.turns:
            lines.append(f"### 第 {turn.turn_number} 轮\n")
            lines.append(f"> **用户**: {turn.user}\n")
            lines.append(f"> **客服**: {turn.assistant}\n")

        # 八、总结
        lines.append("## 八、总结\n")
        lines.append(eval_result.summary + "\n")

        return "\n".join(lines)

    def generate_json(
        self,
        dialog_result: DialogResult,
        eval_result: EvalResult,
        rubric: TaskRubric,
        task_instruction: str,
    ) -> dict:
        """生成 JSON 格式评测数据"""
        return {
            "meta": {
                "task_instruction": task_instruction,
                "generated_at": datetime.datetime.now().isoformat(),
            },
            "scenario": dialog_result.scenario.to_dict(),
            "rubric": rubric.to_dict(),
            "dialog": {
                "turns": len(dialog_result.turns),
                "finished_reason": dialog_result.finished_reason,
                "records": [
                    {"turn": t.turn_number, "user": t.user, "assistant": t.assistant}
                    for t in dialog_result.turns
                ],
            },
            "evaluation": {
                "overall_score": eval_result.overall_score,
                "dimensions": {
                    dim["name"]: {
                        "score": eval_result.dimensions.get(dim["key"], DimensionScore(0, "")).score,
                        "reason": eval_result.dimensions.get(dim["key"], DimensionScore(0, "")).reason,
                    }
                    for dim in DIMENSIONS
                },
                "violations": eval_result.violations,
                "good_points": eval_result.good_points,
                "summary": eval_result.summary,
            },
        }

    def save_report(
        self,
        dialog_result: DialogResult,
        eval_result: EvalResult,
        rubric: TaskRubric,
        task_instruction: str,
        scenario_name: str,
    ) -> Path:
        """保存评测报告到文件"""
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = scenario_name.replace(" ", "_")
        filename = f"report_{safe_name}_{ts}.md"
        report_path = settings.output_dir / "reports" / filename

        md = self.generate_markdown(
            dialog_result, eval_result, rubric, task_instruction
        )
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(md, encoding="utf-8")

        # 同时保存 JSON
        json_filename = f"report_{safe_name}_{ts}.json"
        json_path = settings.output_dir / "reports" / json_filename
        data = self.generate_json(
            dialog_result, eval_result, rubric, task_instruction
        )
        json_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        return report_path

    def generate_pdf(
        self,
        dialog_result: DialogResult,
        eval_result: EvalResult,
        rubric: TaskRubric,
        task_instruction: str,
    ) -> BytesIO:
        """生成 PDF 格式评测报告（返回 BytesIO 供下载）"""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()
        story = []

        # 自定义样式
        title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=18, spaceAfter=10)
        h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=14, spaceAfter=6, spaceBefore=12)
        body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=10, leading=14)
        small_style = ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, textColor=HexColor('#666'))

        # 标题
        story.append(Paragraph("EvalAgent 多轮对话评测报告", title_style))
        story.append(Paragraph(
            f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
            small_style
        ))
        story.append(Spacer(1, 8*mm))

        # 一、任务概述
        story.append(Paragraph("一、任务概述", h2_style))
        story.append(Paragraph(f"任务目标: {rubric.task_goal}", body_style))
        if rubric.role:
            story.append(Paragraph(f"AI 角色: {rubric.role}", body_style))
        story.append(Spacer(1, 4*mm))

        # 二、评测场景
        story.append(Paragraph("二、评测场景", h2_style))
        s = dialog_result.scenario
        story.append(Paragraph(f"用户画像: {s.persona_name}  |  轮次: {len(dialog_result.turns)}  |  原因: {self._end_reason_text(dialog_result.finished_reason)}", body_style))
        story.append(Spacer(1, 6*mm))

        # 三、评分
        story.append(Paragraph("三、综合评分", h2_style))
        score = eval_result.overall_score
        color = "#43a047" if score >= 80 else ("#ef6c00" if score >= 60 else "#c62828")
        story.append(Paragraph(f"<font color='{color}'><b>总分: {score}/100</b></font>", body_style))
        story.append(Spacer(1, 4*mm))

        # 维度表格
        story.append(Paragraph("各维度评分:", h2_style))
        table_data = [["维度", "权重", "得分", "评分理由"]]
        for dim in DIMENSIONS:
            key = dim["key"]
            if key in eval_result.dimensions:
                ds = eval_result.dimensions[key]
                table_data.append([
                    dim["name"],
                    f"{dim['weight']*100:.0f}%",
                    f"{ds.score}/5",
                    ds.reason[:60]
                ])
        t = Table(table_data, colWidths=[70, 40, 40, 350])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#ffffff')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#ccc')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(t)
        story.append(Spacer(1, 6*mm))

        # 四、违规
        if eval_result.violations:
            story.append(Paragraph("四、违规项", h2_style))
            for v in eval_result.violations:
                story.append(Paragraph(f"• {v}", body_style))
            story.append(Spacer(1, 4*mm))

        # 五、总结
        if eval_result.summary:
            story.append(Paragraph("五、总结", h2_style))
            story.append(Paragraph(eval_result.summary, body_style))

        doc.build(story)
        buf.seek(0)
        return buf

    def generate_excel(self, results: list[dict]) -> BytesIO:
        """批量生成 Excel 格式评测报告"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "评测结果"

        headers = ["场景", "总分", *[d["name"] for d in DIMENSIONS],
                     "违规数", "轮次", "结束原因"]

        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        ha = Alignment(horizontal="center", vertical="center")

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = ha

        for idx, r in enumerate(results, 2):
            s = r["scenario"]; e = r["evaluation"]; dr = r["dialog"]
            ws.cell(row=idx, column=1, value=s.persona_name)
            ws.cell(row=idx, column=2, value=e.overall_score)
            for di, dim in enumerate(DIMENSIONS):
                key = dim["key"]
                ws.cell(row=idx, column=3+di,
                        value=e.dimensions[key].score if key in e.dimensions else "")
            ws.cell(row=idx, column=3+len(DIMENSIONS), value=len(e.violations))
            ws.cell(row=idx, column=4+len(DIMENSIONS), value=len(dr.turns))
            ws.cell(row=idx, column=5+len(DIMENSIONS),
                    value=getattr(dr, 'finished_reason', ''))

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = \
                min(max(len(str(c.value or "")) for c in col) + 2, 25)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _end_reason_text(self, reason: str) -> str:
        mapping = {
            "max_turns": "达到最大轮次限制",
            "end_signal": "用户模拟器发送结束信号",
            "natural_end": "对话自然结束",
            "error": "执行出错",
        }
        return mapping.get(reason, reason)
